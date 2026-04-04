"""Reports routes — Excel (.xlsx) export"""
import io
import json
from flask import Blueprint, send_file
from flask_jwt_extended import jwt_required
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from models import Document

reports_bp = Blueprint('reports', __name__)


@reports_bp.route('/export-xlsx', methods=['GET'])
@jwt_required()
def export_xlsx():
    """Export monitoring log as RECORDS INCOMING LOGBOOK Excel file matching excel.pdf layout"""
    docs = Document.query.order_by(Document.date_received.asc(), Document.time_received.asc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = 'RECORDS INCOMING LOGBOOK'

    # Page setup for landscape
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.paperSize = ws.PAPERSIZE_LEGAL

    # Styles
    header_font = Font(name='Arial', bold=True, size=11)
    title_font = Font(name='Arial', bold=True, size=14)
    cell_font = Font(name='Arial', size=9)
    header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    summary_fill = PatternFill(start_color='FEF9C3', end_color='FEF9C3', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='top', wrap_text=True)

    # Title row
    ws.merge_cells('A1:H1')
    ws['A1'] = 'RECORDS INCOMING LOGBOOK'
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='center')

    # Column headers (row 3)
    headers = ['#', 'DATE / CONTROL-REF NO.', 'FROM', 'SUBJECT', 'TO OPM', 'IN', 'OUT', 'DURATION']
    col_widths = [5, 20, 25, 40, 20, 10, 10, 10]

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
        ws.column_dimensions[chr(64 + col_idx)].width = width

    # Data rows
    for i, doc in enumerate(docs, 1):
        row = i + 3
        target_divs = json.loads(doc.target_divisions) if doc.target_divisions else []
        opm_text = ', '.join(target_divs) if target_divs else (doc.target_division or '')

        values = [
            i,
            f"{doc.tracking_number}\n{doc.time_received or ''}",
            f"{doc.sender or ''}\n{doc.sender_address or ''}",
            doc.subject or '',
            f"OPM: {opm_text}\nDate: {doc.date_received or ''}\nTime: {doc.time_received or ''}",
            doc.time_received or '',
            '',  # OUT — filled manually
            '',  # DURATION — filled manually
        ]

        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.font = cell_font
            cell.alignment = left_align if col_idx in (3, 4, 5) else center_align
            cell.border = thin_border

    # Summary row
    summary_row = len(docs) + 5
    ws.merge_cells(f'A{summary_row}:D{summary_row}')
    summary_cell = ws.cell(row=summary_row, column=1, value=f'# of Transactions: {len(docs)}')
    summary_cell.font = Font(name='Arial', bold=True, size=10)
    summary_cell.fill = summary_fill

    # Write to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    from datetime import date
    filename = f"PPA_RECORDS_INCOMING_LOGBOOK_{date.today().isoformat()}.xlsx"
    return send_file(output, as_attachment=True, download_name=filename,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
